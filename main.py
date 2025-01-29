import pandas as pd
from bs4 import BeautifulSoup
import logging
from datetime import datetime
import os
from typing import Dict, List, Optional, Tuple
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'partner_scraping_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)


class PartnerScraper:
    def __init__(self):
        self.required_fields = {
            "Number": str,
            "Siemens Batch": str,
            "Name": str,
            "Partner Batch": str,
            "Locations": str,
            "Office address": str,
            "Contact Name": str,
            "Contact Email": str,
            "Contact Telephone": str,
            "Contact website": str,
            "Proced Specialization": str
        }

    def clean_text(self, text: str) -> str:
        """Clean and normalize text content"""
        if not text:
            return ""
        # Remove extra whitespace and normalize newlines
        text = re.sub(r'\s+', ' ', text.strip())
        # Remove special characters but keep basic punctuation and common symbols
        text = re.sub(r'[^\w\s.,;?!@\-_+()/#]', '', text)
        return text

    def validate_email(self, email: str) -> str:
        """Validate and clean email addresses"""
        if not email:
            return ""
        email = email.lower().strip()
        # Basic email validation
        if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            return email
        logging.warning(f"Invalid email format found: {email}")
        return ""

    def extract_address_components(self, address_td: BeautifulSoup) -> Tuple[str, str, str, str]:
        """
        Extract address components from the address TD element
        Returns: Tuple of (street, city, state/region, country)
        """
        if not address_td:
            return ("", "", "", "")

        try:
            # Find the specific address span within the correct column
            address_span = address_td.find("span", class_="pl-results-td-contact-plocez__Address__c")
            if not address_span:
                return ("", "", "", "")

            # Get all text elements, removing empty strings and extra whitespace
            address_parts = [part.strip() for part in address_span.stripped_strings if part.strip()]

            # Initialize components
            street = city = state = country = ""

            if len(address_parts) >= 1:
                street = address_parts[0]
            if len(address_parts) >= 2:
                city = address_parts[1]
            if len(address_parts) >= 3:
                # Check if the third part contains both state and country
                state_country = address_parts[2].split(',')
                if len(state_country) == 2:
                    state = state_country[0].strip()
                    country = state_country[1].strip()
                else:
                    state = address_parts[2]
            if len(address_parts) >= 4 and not country:
                country = address_parts[3]

            return (
                self.clean_text(street),
                self.clean_text(city),
                self.clean_text(state),
                self.clean_text(country)
            )
        except Exception as e:
            logging.error(f"Error extracting address components: {str(e)}")
            return ("", "", "", "")

    def parse_partner_row(self, html_snippet: str) -> Dict:
        """Parse a single partner row with enhanced error handling and validation"""
        try:
            soup = BeautifulSoup(html_snippet, "html.parser")

            # Number
            number_text = ""
            number_tag = soup.find("span", class_="pl-results-row-no")
            number_text = self.clean_text(number_tag.get_text()) if number_tag else ""

            # Siemens Batch
            siemens_batch = ""
            hidden_td = soup.find("td", class_="sf-hidden")
            if hidden_td:
                hidden_text = hidden_td.get_text(strip=True)
                try:
                    score_match = re.search(r'sortScore\s*:\s*(\d+\.?\d*)', hidden_text)
                    if score_match:
                        siemens_batch = score_match.group(1)
                except Exception as e:
                    logging.warning(f"Error parsing Siemens Batch: {str(e)}")

            # Partner Name
            partner_name = ""
            name_tag = soup.find("a", class_="pl-results-partner-name")
            partner_name = self.clean_text(name_tag.get_text()) if name_tag else ""

            # Partner Batch with enhanced parsing
            partner_batch = ""
            partner_type_span = soup.find("span", id=lambda x: x and "resultsPartnerType" in x)
            if partner_type_span:
                batch_items = partner_type_span.find_all("li", class_="list-group-item")
                partner_batch = ", ".join(self.clean_text(item.get_text()) for item in batch_items)

            # Locations with number validation
            locations_text = ""
            partner_info_span = soup.find("span", class_="pl-results-partner-info")
            if partner_info_span:
                count_span = partner_info_span.find("span", class_="pl-results-partner-count")
                if count_span:
                    text = count_span.get_text(strip=True)
                    locations_match = re.search(r'Locations:\s*(\d+)', text)
                    if locations_match:
                        locations_text = locations_match.group(1)

            # Office address with formatting
            office_address = ""
            # Find the specific address column using proper selector
            #address_td = soup.find("td", class_="pl-results-td-address")
            address_span = soup.find("span", class_="pl-results-td-address-plocez__Mailing_Address__c")
            if address_span:
                address_detail = address_span.find("span", class_="pl-results-value")
                if address_detail:
                    office_address = self.clean_text(address_span.get_text())

            # Contact information with validation
            contact_name = ""
            contact_name_span = soup.find("span", class_="pl-results-td-contact-plocez__Contact__c")
            if contact_name_span:
                    contact_detail = contact_name_span.find("span", class_="pl-results-value")
                    if contact_detail:
                        contact_name = self.clean_text(contact_detail.get_text())

            # Email with validation
            contact_email = ""
            email_span = soup.find("span", class_="pl-results-td-contact-plocez__Email__c")
            if email_span:
                mailto = email_span.find("a", href=True)
                if mailto:
                    contact_email = self.validate_email(mailto.get_text())

            # Phone with validation
            contact_phone = ""
            phone_span = soup.find("span", class_="pl-results-td-contact-plocez__Phone__c")
            if phone_span:
                phone_link = phone_span.find("a", href=True)
                if phone_link:
                    contact_phone = self.clean_text(phone_link.get_text())

            # Website with validation
            contact_website = ""
            website_span = soup.find("span", class_="pl-results-td-contact-PLP_Website__c")
            if website_span:
                web_link = website_span.find("a", href=True)
                if web_link:
                    href = web_link.get('href', '')
                    contact_website = href if href.startswith(('http://', 'https://')) else ''

            # Specializations with deduplication
            specializations = set()
            spec_container = soup.find("span", class_="pl-results-td-account-Product_Specialization__c")
            if spec_container:
                for li in spec_container.find_all("li", class_="list-group-item"):
                    spec = self.clean_text(li.get_text())
                    if spec:
                        specializations.add(spec)
            proced_specialization = ", ".join(sorted(specializations))

            return {
                "Number": number_text,
                "Siemens Batch": siemens_batch,
                "Name": partner_name,
                "Partner Batch": partner_batch,
                "Locations": locations_text,
                "Office address": office_address,
                "Contact Name": contact_name,
                "Contact Email": contact_email,
                "Contact Telephone": contact_phone,
                "Contact website": contact_website,
                "Proced Specialization": proced_specialization
            }

        except Exception as e:
            logging.error(f"Error parsing partner row: {str(e)}")
            return {field: "" for field in self.required_fields}

    def parse_all_partners(self, html_content: str) -> List[Dict]:
        """Parse all partner rows from the HTML content"""
        try:
            soup = BeautifulSoup(html_content, "html.parser")
            partner_rows = soup.find_all("td", class_="pl-results-td-row-no")

            results = []
            #for idx, row in enumerate(partner_rows, 1):
            for idx, row_num_td in enumerate(partner_rows, 1):
                try:
                    # 2) Each "row_num_td" is a <td>.  The <tr> is its parent:
                    parent_tr = row_num_td.find_parent("tr")

                    # 3) Convert that <tr> to a string and parse it:
                    row_html = str(parent_tr)
                    row_data = self.parse_partner_row(row_html)

                    results.append(row_data)
                    logging.info(
                        f"Processed partner {idx}/{len(partner_rows)}: {row_data.get('Name', '(no name)')}"
                    )
                    #row_html = str(row)
                    #row_data = self.parse_partner_row(row_html)
                    #results.append(row_data)
                    #logging.info(f"Processed partner {idx}/{len(partner_rows)}: {row_data['Name']}")
                except Exception as e:
                    logging.error(f"Error processing row {idx}: {str(e)}")

            return results
        except Exception as e:
            logging.error(f"Error parsing HTML content: {str(e)}")
            return []

    def format_excel(self, workbook, df: pd.DataFrame) -> None:
        """Apply formatting to Excel workbook"""
        worksheet = workbook.active

        # Format header
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Adjust column widths
        for col in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col)
            max_length = max(
                len(str(df.columns[col - 1])),
                df.iloc[:, col - 1].astype(str).map(len).max()
            )
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Format data cells
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(vertical='center', wrap_text=True)


def main():
    # Input and output file paths
    input_file = "Partners_Mendix.htm"
    output_dir = "output"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    try:
        # Read HTML content
        with open(input_file, "r", encoding="utf-8") as f:
            html_content = f.read()

        # Initialize scraper and process data
        scraper = PartnerScraper()
        partner_data = scraper.parse_all_partners(html_content)

        if not partner_data:
            logging.error("No partner data found!")
            return

        # Convert to DataFrame
        df = pd.DataFrame(partner_data)

        # Save to Excel with formatting
        output_file = os.path.join(output_dir, f"partner_data_{timestamp}.xlsx")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            scraper.format_excel(writer.book, df)

        # Generate summary
        summary = {
            "Total Partners": len(df),
            "Unique Batches": df["Partner Batch"].nunique(),
            "Partners with Email": df["Contact Email"].notna().sum(),
            "Partners with Phone": df["Contact Telephone"].notna().sum(),
            "Unique Specializations": len(set(s.strip() for specs in df["Proced Specialization"].dropna()
                                              for s in specs.split(",")))
        }

        # Save summary
        summary_file = os.path.join(output_dir, f"summary_{timestamp}.txt")
        with open(summary_file, "w") as f:
            for key, value in summary.items():
                f.write(f"{key}: {value}\n")

        logging.info(f"Successfully exported {len(df)} partners to {output_file}")

    except Exception as e:
        logging.error(f"Script failed: {str(e)}")
        raise


if __name__ == "__main__":
    main()